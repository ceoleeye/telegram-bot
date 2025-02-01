import os
import sys
import json
import requests
import re
from datetime import datetime, timedelta, time
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import (
    Application,
    MessageHandler,
    filters,
    CommandHandler,
    CallbackContext,
    JobQueue
)

# 한국 시간대를 위한 모듈
from zoneinfo import ZoneInfo


# ======================================================================
# 1) 설정값 (엑셀, 텔레그램 봇, 파일 경로 등)
# ======================================================================
DOWNLOAD_FOLDER = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILENAME = "근무일정표 mina perhonen DDP 2024.xlsx"  # 실제 파일명
EXCEL_FILE_PATH = os.path.join(DOWNLOAD_FOLDER, EXCEL_FILENAME)

SHEET_ID = "1Sn9_VmyQ9o067QHwKmP59hKXQHcnYIMhUJMOWHJ3hPA"  # 실제 구글 시트 ID
GOOGLE_SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

BOT_TOKEN = "7653968457:AAEuaXC-QbG0dOE9LkoAEC2xtqX-P9V7rXA"  # 실제 봇 토큰
CHAT_ID = -1002410880873  # 단체(그룹) Chat ID (음수 가능)

TASKS_JSON = "tasks.json"
tasks_store = {}  # 예: { "강혜경": ["업무1", "업무2"], ... }

# 직급순 (업무 정렬)
FIXED_ORDER = ["강혜경", "이예찬", "최영윤", "김민성", "유민아"]

# 요일 영문 -> 한글
weekday_map = {
    "Mon": "월",
    "Tue": "화",
    "Wed": "수",
    "Thu": "목",
    "Fri": "금",
    "Sat": "토",
    "Sun": "일",
}

# 엑셀에서 날짜를 찾을 때,
# row+1=오전 근무자, row+2=오후, row+3=휴무, row+4=특이사항
SHIFT_DEFS = [
    ("11:00", 5),
    ("12:00", 6),
    ("13:00", 7),
    ("15:00", 8),
    ("16:30", 9),
    ("18:00", 10),
    ("19:00", 11),
]


# ======================================================================
# 2) 엑셀 다운로드/삭제
# ======================================================================
def download_excel_file():
    print("[진행] 구글 스프레드시트 XLSX 다운로드 중...")
    resp = requests.get(GOOGLE_SHEET_URL)
    with open(EXCEL_FILE_PATH, "wb") as f:
        f.write(resp.content)
    print(f"[진행] 다운로드 완료: {EXCEL_FILE_PATH}")


def remove_excel_file():
    if os.path.exists(EXCEL_FILE_PATH):
        os.remove(EXCEL_FILE_PATH)
        print(f"[진행] 엑셀 파일 삭제 완료: {EXCEL_FILE_PATH}")
    else:
        print("[진행] 엑셀 파일이 이미 없거나 삭제된 상태.")


# ======================================================================
# 3) 날짜 셀 찾기 & 근무 정보
# ======================================================================
def find_date_cell(sheet, target_date: datetime):
    yyyymmdd = target_date.strftime("%Y-%m-%d")
    if os.name == 'nt':
        mmdd = target_date.strftime("%#m/%#d")
        mmdd_kor = target_date.strftime("%#m/%#d (%a)")
    else:
        mmdd = target_date.strftime("%-m/%-d")
        mmdd_kor = target_date.strftime("%-m/%-d (%a)")

    for row in sheet.iter_rows():
        for cell in row:
            val = str(cell.value).strip() if cell.value else ""
            if (yyyymmdd in val) or (mmdd in val) or (mmdd_kor in val):
                print(f"[진행] 날짜 셀 발견: row={cell.row}, col={cell.column}, 값={cell.value}")
                return cell.row, cell.column
    print("[주의] 해당 날짜를 엑셀에서 찾지 못했습니다.")
    return None, None


def get_work_info(sheet, date_row, date_col):
    morning = sheet.cell(row=date_row+1, column=date_col).value or ""
    afternoon = sheet.cell(row=date_row+2, column=date_col).value or ""
    off_info = sheet.cell(row=date_row+3, column=date_col).value or ""
    special = sheet.cell(row=date_row+4, column=date_col).value or ""
    return str(morning).strip(), str(afternoon).strip(), str(off_info).strip(), str(special).strip()


# ======================================================================
# 4) tasks.json 로드/저장
# ======================================================================
def load_tasks():
    global tasks_store
    if os.path.exists(TASKS_JSON):
        with open(TASKS_JSON, "r", encoding="utf-8") as f:
            tasks_store = json.load(f)
        print("[진행] tasks.json 로드 완료.")
    else:
        tasks_store = {}
        print("[진행] tasks.json 파일이 없어 새로 생성 예정.")


def save_tasks():
    with open(TASKS_JSON, "w", encoding="utf-8") as f:
        json.dump(tasks_store, f, ensure_ascii=False, indent=2)
    print("[진행] tasks.json 저장 완료.")


# ======================================================================
# 5) /closing -> 내일 근무자 보고
# ======================================================================
async def handle_closing(update: Update, context: CallbackContext):
    """
    /closing => 내일 날짜 보고 + 특이사항
    """
    chat_id = update.effective_chat.id

    # "내일"을 한국 시간으로 계산
    tomorrow = datetime.now(ZoneInfo("Asia/Seoul")) + timedelta(days=1)

    download_excel_file()
    wb = load_workbook(EXCEL_FILE_PATH)
    sheet = wb.active

    rowcol = find_date_cell(sheet, tomorrow)
    if not rowcol or not rowcol[0]:
        remove_excel_file()
        await context.bot.send_message(chat_id=chat_id,
            text="내일 날짜를 엑셀에서 찾지 못했습니다.")
        return

    date_row, date_col = rowcol
    morning, afternoon, off_info, special = get_work_info(sheet, date_row, date_col)
    remove_excel_file()

    if os.name == 'nt':
        mm = tomorrow.strftime("%#m")
        dd = tomorrow.strftime("%#d")
    else:
        mm = tomorrow.strftime("%-m")
        dd = tomorrow.strftime("%-d")
    eng_wd = tomorrow.strftime("%a")

    lines = []
    lines.append("전시장 마감했습니다.")
    lines.append("")
    lines.append(f"{mm}/{dd} ({weekday_map.get(eng_wd, eng_wd)})")
    lines.append(f"오전: {morning}")
    lines.append(f"오후: {afternoon}")
    lines.append(f"휴무: {off_info}")
    lines.append("")
    lines.append("[특이사항]")
    if special.strip():
        lines.append(special)
    else:
        lines.append("없습니다.")

    await context.bot.send_message(chat_id=chat_id, text="\n".join(lines))


# ======================================================================
# 6) /opening -> 오늘 근무자 + 업무 + 특이사항
# ======================================================================
async def handle_opening(update: Update, context: CallbackContext):
    """
    /opening => 오늘 날짜 근무자+특이사항+업무
    """
    chat_id = update.effective_chat.id
    try:
        download_excel_file()
        wb = load_workbook(EXCEL_FILE_PATH)
        sheet = wb.active

        # "오늘"을 한국 시간으로 계산
        today = datetime.now(ZoneInfo("Asia/Seoul"))
        rowcol = find_date_cell(sheet, today)
        if not rowcol or not rowcol[0]:
            remove_excel_file()
            await context.bot.send_message(chat_id=chat_id,
                text="오늘 날짜를 엑셀에서 찾지 못했습니다.")
            return

        date_row, date_col = rowcol
        morning, afternoon, off_info, special = get_work_info(sheet, date_row, date_col)
        remove_excel_file()

        if os.name == 'nt':
            mm = today.strftime("%#m")
            dd = today.strftime("%#d")
        else:
            mm = today.strftime("%-m")
            dd = today.strftime("%-d")
        wd = today.strftime("%a")
        kor_wd = weekday_map.get(wd, wd)

        lines = []
        lines.append(f"{mm}/{dd} ({kor_wd})")
        lines.append("[금일 근무자]")
        lines.append(f"오전: {morning}")
        lines.append(f"오후: {afternoon}")
        lines.append(f"휴무: {off_info}")
        lines.append("")
        lines.append("[금일 업무]")

        any_person = False
        for person in FIXED_ORDER:
            if person in tasks_store and tasks_store[person]:
                lines.append(person)
                for t in tasks_store[person]:
                    lines.append(t)
                lines.append("")
                any_person = True

        if not any_person:
            lines.append("등록된 업무가 없습니다.\n")

        lines.append("[특이사항]")
        if special.strip():
            lines.append(special)
        else:
            lines.append("없습니다.")

        await context.bot.send_message(chat_id=chat_id,
            text="\n".join(lines))

    except Exception as e:
        err_msg = f"/opening 오류: {e}"
        await context.bot.send_message(chat_id=chat_id, text=err_msg)


# ======================================================================
# 7) /today -> 오늘 지원 시간표
# ======================================================================
async def handle_today(update: Update, context: CallbackContext):
    chat_id = update.effective_chat.id
    try:
        download_excel_file()
        wb = load_workbook(EXCEL_FILE_PATH)
        sheet = wb.active

        # "오늘"을 한국 시간으로
        nowdt = datetime.now(ZoneInfo("Asia/Seoul"))
        rowcol = find_date_cell(sheet, nowdt)
        if not rowcol or not rowcol[0]:
            remove_excel_file()
            await context.bot.send_message(chat_id=chat_id,
                text="오늘 날짜 셀을 찾지 못했습니다.")
            return

        date_row, date_col = rowcol
        if os.name == 'nt':
            mm = nowdt.strftime("%#m")
            dd = nowdt.strftime("%#d")
        else:
            mm = nowdt.strftime("%-m")
            dd = nowdt.strftime("%-d")

        lines = [f"[{mm}/{dd} 지원 시간표]"]
        for (time_str, offset) in SHIFT_DEFS:
            r = date_row + offset
            staff_list = []
            v1 = sheet.cell(row=r, column=date_col).value
            v2 = sheet.cell(row=r, column=date_col+1).value

            if v1: staff_list.append(str(v1).strip())
            if v2: staff_list.append(str(v2).strip())
            if not staff_list:
                staff_list.append("(근무자 없음)")

            lines.append(f"{time_str} ➜ {', '.join(staff_list)}")

        await context.bot.send_message(chat_id=chat_id, text="\n".join(lines))

    except Exception as e:
        err_msg = f"/today 오류: {e}"
        await context.bot.send_message(chat_id=chat_id, text=err_msg)
    finally:
        remove_excel_file()


# ======================================================================
# 8) /강혜경, /이예찬 등 (업무 등록)
# ======================================================================
async def handle_person_command(update: Update, context: CallbackContext, person_name: str):
    chat_id = update.effective_chat.id
    txt = update.message.text
    lines = txt.split("\n")
    if len(lines) <= 1:
        await context.bot.send_message(chat_id=chat_id,
            text=f"{person_name} 님의 업무를 한 줄씩 입력해주세요.")
        return
    tasks = [l.strip() for l in lines[1:] if l.strip()]
    if not tasks:
        await context.bot.send_message(chat_id=chat_id,
            text=f"{person_name} 님의 업무가 비어있습니다.")
        return

    if person_name not in tasks_store:
        tasks_store[person_name] = []
    tasks_store[person_name].extend(tasks)
    save_tasks()

    await context.bot.send_message(chat_id=chat_id,
        text=f"[{person_name}] 님의 업무가 추가되었습니다!")


async def handle_kang(update, context): await handle_person_command(update, context, "강혜경")
async def handle_lee(update, context):  await handle_person_command(update, context, "이예찬")
async def handle_choi(update, context): await handle_person_command(update, context, "최영윤")
async def handle_kim(update, context):  await handle_person_command(update, context, "김민성")
async def handle_yu(update, context):   await handle_person_command(update, context, "유민아")


# ======================================================================
# 9) /edit, /reset
# ======================================================================
async def handle_edit(update: Update, context: CallbackContext):
    """
    /edit 이예찬
    (새 업무 1)
    (새 업무 2)
    => 덮어쓰기
    """
    chat_id = update.effective_chat.id
    txt = update.message.text
    lines = txt.split("\n")
    if len(lines) <= 1:
        await context.bot.send_message(chat_id=chat_id,
            text="이름과 새 업무 내용을 한 줄씩 입력.\n예) /edit 이예찬\n(업무1)\n(업무2)")
        return

    name_line = lines[0].partition(" ")[2].strip()  # /edit 뒤
    if not name_line:
        await context.bot.send_message(chat_id=chat_id,
            text="이름이 누락됨. 예) /edit 이예찬\n(업무1)\n(업무2)")
        return

    name = name_line.split("\n")[0].strip()
    tasks = [l.strip() for l in lines[1:] if l.strip()]
    if not tasks:
        await context.bot.send_message(chat_id=chat_id,
            text="새 업무 내용이 비어있습니다.")
        return

    tasks_store[name] = tasks
    save_tasks()

    await context.bot.send_message(chat_id=chat_id,
        text=f"[{name}]의 업무를 새로 설정했습니다!")


async def handle_reset(update: Update, context: CallbackContext):
    """
    /reset => 전체 업무 리셋
    """
    chat_id = update.effective_chat.id
    tasks_store.clear()
    if os.path.exists(TASKS_JSON):
        os.remove(TASKS_JSON)
    await context.bot.send_message(chat_id=chat_id,
        text="(전체) 모든 업무가 초기화되었습니다!")


# ======================================================================
# 10) 매일 20:00 => /closing + /reset (자동)
# ======================================================================
async def scheduled_closing_and_reset(context: CallbackContext):
    """
    매일 20:00 => /closing 로직 + 전체 업무 리셋
    """
    chat_id = CHAT_ID

    # 1) /closing
    tomorrow = datetime.now(ZoneInfo("Asia/Seoul")) + timedelta(days=1)
    download_excel_file()
    wb = load_workbook(EXCEL_FILE_PATH)
    sheet = wb.active

    rowcol = find_date_cell(sheet, tomorrow)
    if rowcol and rowcol[0]:
        date_row, date_col = rowcol
        morning, afternoon, off_info, special = get_work_info(sheet, date_row, date_col)

        if os.name == 'nt':
            mm = tomorrow.strftime("%#m")
            dd = tomorrow.strftime("%#d")
        else:
            mm = tomorrow.strftime("%-m")
            dd = tomorrow.strftime("%-d")
        eng_wd = tomorrow.strftime("%a")

        lines = []
        lines.append("전시장 마감했습니다.")
        lines.append("")
        lines.append(f"{mm}/{dd} ({weekday_map.get(eng_wd, eng_wd)})")
        lines.append(f"오전: {morning}")
        lines.append(f"오후: {afternoon}")
        lines.append(f"휴무: {off_info}")
        lines.append("")
        lines.append("[특이사항]")
        if special.strip():
            lines.append(special)
        else:
            lines.append("없습니다.")

        await context.bot.send_message(chat_id=chat_id, text="\n".join(lines))
    else:
        await context.bot.send_message(chat_id=chat_id,
            text="(자동) 내일 날짜를 엑셀에서 찾지 못했습니다.")

    remove_excel_file()

    # 2) /reset
    tasks_store.clear()
    if os.path.exists(TASKS_JSON):
        os.remove(TASKS_JSON)

    await context.bot.send_message(
        chat_id=chat_id,
        text="(20시 자동) 모든 업무가 초기화되었습니다!"
    )


# ======================================================================
# 11) 시프트별 알림
# ======================================================================
async def scheduled_shift_notify(context: CallbackContext):
    shift_time_str = context.job.data
    try:
        download_excel_file()
        wb = load_workbook(EXCEL_FILE_PATH)
        sheet = wb.active
        today = datetime.now(ZoneInfo("Asia/Seoul"))
        rowcol = find_date_cell(sheet, today)
        if not rowcol or not rowcol[0]:
            remove_excel_file()
            await context.bot.send_message(chat_id=CHAT_ID,
                text=f"{shift_time_str} 알림 실패 - 오늘 날짜 찾기 실패.")
            return

        date_row, date_col = rowcol
        offset = None
        for (ts, off) in SHIFT_DEFS:
            if ts == shift_time_str:
                offset = off
                break

        lines = []
        if offset is None:
            lines.append(f"{shift_time_str}는 정의되지 않은 시프트.")
        else:
            row_i = date_row + offset
            main_val = sheet.cell(row=row_i, column=date_col).value
            sub_val = sheet.cell(row=row_i, column=date_col+1).value
            staff_list = []
            if main_val: staff_list.append(str(main_val).strip())
            if sub_val: staff_list.append(str(sub_val).strip())
            if not staff_list:
                staff_list.append("(근무자 없음)")

            lines.append(f"{shift_time_str} => {', '.join(staff_list)}")

        remove_excel_file()
        await context.bot.send_message(chat_id=CHAT_ID,
            text="\n".join(lines))

    except Exception as e:
        await context.bot.send_message(chat_id=CHAT_ID,
            text=f"{shift_time_str} 알림 오류: {e}")


# ======================================================================
# 12) fallback_command: /YYYYMMDD => 특정 날짜 (숫자없으면 내일)
# ======================================================================
async def fallback_command(update: Update, context: CallbackContext):
    """
    /YYYYMMDD 로 입력한 경우 -> 8자리 숫자만 있으면 그 날짜,
                               숫자가 없거나 틀리면 => 내일 날짜
    그 외 명령어는 알수없음?
    """
    chat_id = update.effective_chat.id
    cmd = update.message.text.lstrip("/")  # 예) YYYYMMDD

    # 1) 8자리 숫자인지
    if re.fullmatch(r"\d{8}", cmd):
        # => 특정 날짜
        try:
            dt = datetime.strptime(cmd, "%Y%m%d")
        except ValueError:
            await context.bot.send_message(chat_id=chat_id,
                text="날짜형식이 잘못됨. 예) /20250201")
            return
    else:
        # => 내일 (한국 시간 + 1일)
        dt = datetime.now(ZoneInfo("Asia/Seoul")) + timedelta(days=1)

    # 이제 dt 날짜 출력
    try:
        download_excel_file()
        wb = load_workbook(EXCEL_FILE_PATH)
        sheet = wb.active
        rowcol = find_date_cell(sheet, dt)
        if not rowcol or not rowcol[0]:
            remove_excel_file()
            if re.fullmatch(r"\d{8}", cmd):
                await context.bot.send_message(chat_id=chat_id,
                    text=f"{cmd} => 엑셀에서 못 찾음.")
            else:
                await context.bot.send_message(chat_id=chat_id,
                    text="내일 날짜를 엑셀에서 찾지 못했습니다.")
            return

        date_row, date_col = rowcol
        morning, afternoon, off_info, special = get_work_info(sheet, date_row, date_col)
        remove_excel_file()

        if os.name == 'nt':
            mm = dt.strftime("%#m")
            dd = dt.strftime("%#d")
        else:
            mm = dt.strftime("%-m")
            dd = dt.strftime("%-d")
        wd = dt.strftime("%a")
        kor_wd = weekday_map.get(wd, wd)

        lines = []
        lines.append(f"{mm}/{dd} ({kor_wd})")
        lines.append("[금일 근무자]")
        lines.append(f"오전: {morning}")
        lines.append(f"오후: {afternoon}")
        lines.append(f"휴무: {off_info}")
        lines.append("")
        lines.append("[특이사항]")
        lines.append(special if special.strip() else "없음")
        lines.append("")
        lines.append("[지원 시간표]")

        # SHIFT
        for (time_str, offset) in SHIFT_DEFS:
            rowi = date_row + offset
            staff_list = []
            v1 = sheet.cell(row=rowi, column=date_col).value
            v2 = sheet.cell(row=rowi, column=date_col+1).value
            if v1: staff_list.append(str(v1).strip())
            if v2: staff_list.append(str(v2).strip())
            if not staff_list:
                staff_list.append("(근무자 없음)")

            lines.append(f"{time_str} ➜ {', '.join(staff_list)}")

        await context.bot.send_message(chat_id=chat_id, text="\n".join(lines))

    except Exception as e:
        await context.bot.send_message(chat_id=chat_id, text=str(e))


# ======================================================================
# 13) 일반 텍스트 핸들러
# ======================================================================
async def handle_text_message(update: Update, context: CallbackContext):
    text = update.message.text.strip()
    chat_id = update.message.chat_id

    if text == "/":
        usage = (
            "<사용 가능한 명령어 안내> \n\n"
            "🎉 아래 파란 글자를 클릭해도 명령어가 실행됩니다🎉 \n\n"
            "* 오픈/마감 보고 *\n"
            "/opening ➞ 오픈 보고\n"
            "/closing ➞ 마감 보고\n\n"
            "* 업무 등록 관련 *\n"
            "/이름 ➞ 기존 업무 보고에 슬래시만 추가\n"
            "/edit 이름 ➞ 해당 인물 업무 새로 덮어쓰기\n"
            "/reset ➞ 전체 업무 리셋\n\n"
            "* 근무표 관련 *\n"
            "/today ➞ 금일 MD 지원 시간표\n"
            "/YYYYMMDD(실제날짜작성) ➞ 특정 날짜 근무표 + MD 지원\n\n"
            "* 자동기능 *\n"
            " - 매일 09:50, /opening 자동 실행\n"
            " - 매일 20:00, /closing 자동 실행\n"
            " - 매일 20:00, 전체 업무 /reset 자동 실행\n"
            " - 지원 근무 시간 알림 서비스\n"
        )
        await context.bot.send_message(chat_id=chat_id, text=usage)
        return

    print(f"[진행] 일반 텍스트 => 무시: {text}")


# ======================================================================
# main()
# ======================================================================
def main():
    if sys.platform == "win32":
        import asyncio
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

    load_tasks()

    print("[진행] 텔레그램 Application 생성")
    app = Application.builder().token(BOT_TOKEN).build()

    # 명령어(정적)
    app.add_handler(CommandHandler("closing", handle_closing))
    app.add_handler(CommandHandler("opening", handle_opening))
    app.add_handler(CommandHandler("today", handle_today))

    import re
    app.add_handler(MessageHandler(filters.Regex(re.compile(r"^/강혜경(\s|$)")), handle_kang))
    app.add_handler(MessageHandler(filters.Regex(re.compile(r"^/이예찬(\s|$)")), handle_lee))
    app.add_handler(MessageHandler(filters.Regex(re.compile(r"^/최영윤(\s|$)")), handle_choi))
    app.add_handler(MessageHandler(filters.Regex(re.compile(r"^/김민성(\s|$)")), handle_kim))
    app.add_handler(MessageHandler(filters.Regex(re.compile(r"^/유민아(\s|$)")), handle_yu))

    app.add_handler(CommandHandler("edit", handle_edit))
    app.add_handler(CommandHandler("reset", handle_reset))

    # 남은 모든 /슬래시명령 -> fallback_command (/YYYYMMDD)
    app.add_handler(MessageHandler(filters.COMMAND, fallback_command))

    # 일반 텍스트
    app.add_handler(MessageHandler(filters.TEXT, handle_text_message))

    # === 스케줄 ===
    seoul = ZoneInfo("Asia/Seoul")
    job_queue = app.job_queue

    # (1) 매일 9:50 => /opening
    job_queue.run_daily(
        handle_opening,
        time=time(9, 50, tzinfo=seoul)
    )

    # (2) 매일 20:00 => /closing + reset
    job_queue.run_daily(
        scheduled_closing_and_reset,
        time=time(20, 0, tzinfo=seoul)
    )

    # (3) 시프트별 알림
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(10, 55, tzinfo=seoul), data="11:00")
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(11, 55, tzinfo=seoul), data="12:00")
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(12, 55, tzinfo=seoul), data="13:00")
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(14, 55, tzinfo=seoul), data="15:00")
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(16, 25, tzinfo=seoul), data="16:30")
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(17, 55, tzinfo=seoul), data="18:00")
    job_queue.run_daily(callback=scheduled_shift_notify, time=time(18, 55, tzinfo=seoul), data="19:00")

    print("[진행] 봇을 시작합니다! Ctrl+C로 종료하세요.")
    app.run_polling()


if __name__ == "__main__":
    main()
